import os
from zipfile import ZipFile
import pandas as pd
from PyPDF2 import PdfReader
from openpyxl import load_workbook

zipFilesDir = './resources'
zipPath = './tmp/testzip.zip'


def create_zip():
    with ZipFile(zipPath, 'w') as testZip:
        for filename in os.listdir(zipFilesDir):
            filename_with_dir = os.path.join(zipFilesDir, filename)
            filename_full = os.path.abspath(filename_with_dir)
            testZip.write(filename_full, filename)


def delete_zip():
    os.remove(zipPath)


def test_check_zip_file_is_created():
    create_zip()
    assert os.path.isfile(zipPath)


def test_number_of_files_is_same():
    create_zip()
    with ZipFile(zipPath, mode='r') as testZip:
        assert len(testZip.namelist()) == len(os.listdir(zipFilesDir))


def test_files_are_same():
    create_zip()
    with ZipFile(zipPath, 'r') as testZip:
        # check all files from directory exists in zip
        for filename in os.listdir(zipFilesDir):
            assert filename in testZip.namelist()
        # check all files in zip have same length as file in directory
        for zip_item in testZip.filelist:
            filename = os.path.join(os.path.join(zipFilesDir, zip_item.filename))
            zip_size = zip_item.file_size
            file_size = os.path.getsize(filename)
            assert zip_size == file_size


def test_txt():
    create_zip()
    with ZipFile(zipPath, 'r') as testZip:
        assert 'АлтайскийКрай\n' in testZip.read('regions.txt').decode('utf-8')


def test_pdf():
    create_zip()
    with ZipFile(zipPath) as testZip:
        with testZip.open("sprafka.pdf") as pdf_f:
            pdf_file = PdfReader(pdf_f)
            number_of_page = len(pdf_file.pages)
            assert number_of_page == 1
            assert 'Справка о доходах для получения кредита' in pdf_file.pages[0].extract_text()


def test_xls():
    create_zip()
    with ZipFile(zipPath) as testZip:
        with testZip.open("file_example.xls") as xls_f:
            text = pd.read_excel(xls_f).head(2)
            assert 'Hashimoto' in text['Last Name'].values


def test_xlsx():
    create_zip()
    with ZipFile(zipPath) as testZip:
        with testZip.open("file_123.xlsx") as xlsx_f:
            xlsx_file = load_workbook(xlsx_f)
            sheet = xlsx_file.active
            sheet_value = sheet.cell(row=14, column=3).value
            num_rows = sheet.max_row
            num_cols = sheet.max_column
            assert sheet_value == 'Ascencio'
            assert num_rows == 51
            assert num_cols == 8


def test_check_zip_is_removed():
    delete_zip()
    assert not os.path.isfile(zipPath)
